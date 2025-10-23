#!/usr/bin/env python3
"""
Excel Template Generator with Interactive Forms

Creates Excel workbooks with interactive form controls (dropdowns, checkboxes, data validation)
based on CSV field definitions and Python configuration files.

pip install openpyxl

Features:
- Data validation dropdowns
- Checkbox controls  
- Input validation
- Conditional formatting
- Protected sheets with unlocked input cells
"""

import csv
import sys
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
import string

failed_templates = []

def load_template_data(py_file):
    """Load configuration and data from Python file."""
    namespace = {}
    with open(py_file, 'r', encoding='utf-8') as f:
        exec(f.read(), namespace)
    
    config = namespace.get('TEMPLATE_CONFIG', {})
    data = namespace.get('DATA', {})
    return config, data

def load_csv_fields(csv_file):
    """Load field definitions from CSV."""
    fields = []
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            fields.append(row)
    
    fields.sort(key=lambda x: int(x['field_order']))
    return fields

def setup_workbook_styles(wb):
    """Set up custom styles for the workbook."""
    
    # Title style
    title_style = NamedStyle(name="title_style")
    title_style.font = Font(name="Calibri", size=18, bold=True, color="1F4E79")
    title_style.alignment = Alignment(horizontal="center", vertical="center")
    title_style.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    wb.add_named_style(title_style)
    
    # Header style
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_style.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_style.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_style.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    wb.add_named_style(header_style)
    
    # Field label style
    field_label_style = NamedStyle(name="field_label_style")
    field_label_style.font = Font(name="Calibri", size=12, bold=True, color="2F5597")
    field_label_style.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    field_label_style.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    field_label_style.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    wb.add_named_style(field_label_style)
    
    # Input cell style
    input_style = NamedStyle(name="input_style")
    input_style.font = Font(name="Calibri", size=11, color="000000")
    input_style.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    input_style.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    input_style.border = Border(
        left=Side(style="medium", color="4472C4"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    wb.add_named_style(input_style)
    
    # Required field style
    required_style = NamedStyle(name="required_style")
    required_style.font = Font(name="Calibri", size=12, bold=True, color="C00000")
    required_style.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    required_style.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    required_style.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    wb.add_named_style(required_style)
    
    # Description style
    desc_style = NamedStyle(name="desc_style")
    desc_style.font = Font(name="Calibri", size=10, italic=True, color="666666")
    desc_style.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    desc_style.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    wb.add_named_style(desc_style)

def create_overview_sheet(wb, config, fields):
    """Create an overview sheet with template information."""
    ws = wb.create_sheet("Template Overview", 0)
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    
    # Title
    ws.merge_cells('A1:C2')
    ws['A1'] = config['name']
    ws['A1'].style = "title_style"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    
    # Description
    if config.get('description'):
        ws.merge_cells('A4:C4')
        ws['A4'] = config['description']
        ws['A4'].style = "desc_style"
        ws.row_dimensions[4].height = 40
    
    # Template information
    row = 6
    info_data = [
        ("Template Name:", config['name']),
        ("Labels:", config.get('labels', 'N/A')),
        ("Total Fields:", len([f for f in fields if f['field_type'] != 'markdown'])),
        ("Required Fields:", len([f for f in fields if f['required'].lower() == 'true' and f['field_type'] != 'markdown']))
    ]
    
    for label, value in info_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].style = "field_label_style"
        ws[f'B{row}'] = str(value)
        ws[f'B{row}'].style = "input_style"
        row += 1
    
    # Instructions
    row += 2
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "Instructions:"
    ws[f'A{row}'].style = "header_style"
    ws.row_dimensions[row].height = 25
    
    instructions = [
        "• Navigate to the 'Form Fields' tab to complete the template",
        "• Fields marked with (*) are required",
        "• Use dropdown menus for single selections",
        "• Check multiple boxes for multi-select fields",
        "• Hover over cells for additional help text",
        "• The sheet is protected to prevent accidental changes to formatting",
        "• Only input cells (white background) can be edited"
    ]
    
    for instruction in instructions:
        row += 1
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = instruction
        ws[f'A{row}'].style = "desc_style"
        ws.row_dimensions[row].height = 20
    
    return ws

def create_form_sheet(wb, config, fields, data):
    """Create the main form sheet with interactive controls."""
    ws = wb.create_sheet("Form Fields")
    
    # Set column widths
    ws.column_dimensions['A'].width = 3   # Field number
    ws.column_dimensions['B'].width = 35  # Field label
    ws.column_dimensions['C'].width = 50  # Input area
    ws.column_dimensions['D'].width = 30  # Field info/options
    
    # Header row
    headers = ["#", "Field", "Your Response", "Field Information"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.style = "header_style"
    
    ws.row_dimensions[1].height = 25
    
    # Add fields
    current_row = 2
    field_number = 1
    
    for field_def in fields:
        # Merge config defaults into data if needed
        if field_def['field_id'] not in data and field_def['field_id'] in config:
            data[field_def['field_id']] = config[field_def['field_id']]
        
        current_row = add_field_to_sheet(ws, field_def, data, field_number, current_row)
        
        # Only increment counter for non-markdown fields
        if field_def['field_type'] != 'markdown':
            field_number += 1
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    return ws

def add_field_to_sheet(ws, field_def, data, field_number, start_row):
    """Add a single field to the worksheet with appropriate controls."""
    
    field_type = field_def['field_type']
    field_id = field_def['field_id']
    label = field_def['label']
    description = field_def['description']
    data_source = field_def['data_source']
    required = field_def['required'].lower() == 'true'
    placeholder = field_def['placeholder']
    options_type = field_def['options_type']
    default_value = field_def['default_value']
    
    current_row = start_row
    
    # Handle markdown fields as information
    if field_type == 'markdown':
        if description:
            ws.merge_cells(f'A{current_row}:D{current_row}')
            cell = ws[f'A{current_row}']
            # Fix: Move string replacement outside f-string
            formatted_desc = description.replace('\\n', ' ')
            cell.value = f"Information: {formatted_desc}"
            cell.style = "desc_style"
            ws.row_dimensions[current_row].height = 30
            return current_row + 2
        return current_row
    
    # Field number
    ws[f'A{current_row}'] = field_number
    ws[f'A{current_row}'].style = "field_label_style"
    
    # Field label
    label_text = label
    if required:
        label_text += " *"
    ws[f'B{current_row}'] = label_text
    ws[f'B{current_row}'].style = "required_style" if required else "field_label_style"
    
    # Input area (column C)
    input_cell = ws[f'C{current_row}']
    input_cell.style = "input_style"
    
    # Field information (column D)
    info_text = f"Type: {field_type.title()}"
    if field_id:
        info_text += f"\\nID: {field_id}"
    
    if description and description.strip():
        # Fix: Move string replacement outside f-string
        formatted_desc = description.replace('\\n', ' ')
        info_text += f"\\nDescription: {formatted_desc}"
    
    ws[f'D{current_row}'] = info_text
    ws[f'D{current_row}'].style = "desc_style"
    
    # Add interactive controls based on field type
    if field_type in ['dropdown', 'multi-select']:
        add_dropdown_validation(ws, input_cell, field_def, data)
        
        # Add options list to info column
        options = get_field_options(field_def, data)
        if options:
            options_text = "Options:\\n• " + "\\n• ".join(options[:10])
            if len(options) > 10:
                options_text += f"\\n... ({len(options)} total options)"
            ws[f'D{current_row}'].value += f"\\n\\n{options_text}"
            
    elif field_type in ['input', 'textarea']:
        add_text_validation(ws, input_cell, field_def)
        
    elif field_type == 'checkboxes':
        add_checkbox_options(ws, input_cell, field_def, data, current_row)
    
    # Add placeholder/default value
    if placeholder:
        input_cell.comment = Comment(f"Placeholder: {placeholder}", "System")
    elif default_value:
        input_cell.value = default_value
        input_cell.comment = Comment(f"Default value: {default_value}", "System")
    
    # Add help comment
    if description:
        # Fix: Move string replacement outside f-string
        formatted_desc = description.replace('\\n', '\\n')
        help_text = formatted_desc
        instructions = get_filling_instructions(field_type, required, field_def, data)
        if instructions:
            help_text += f"\\n\\nInstructions: {instructions}"
        input_cell.comment = Comment(help_text, "System")
    
    # Set row height
    ws.row_dimensions[current_row].height = 40 if field_type == 'textarea' else 25
    
    # Add spacing
    return current_row + 2

def add_dropdown_validation(ws, cell, field_def, data):
    """Add dropdown data validation to a cell."""
    options = get_field_options(field_def, data)
    
    if options:
        # Create validation list (Excel limit is ~255 characters for formula)
        if len(str(options)) < 200:  # Simple case
            dv = DataValidation(type="list", formula1=f'"{",".join(options)}"')
        else:  # Use a separate range for long lists
            # Create a hidden sheet for long option lists
            if "ValidationData" not in [sheet.title for sheet in ws.parent.worksheets]:
                validation_sheet = ws.parent.create_sheet("ValidationData")
                validation_sheet.sheet_state = 'hidden'
            else:
                validation_sheet = ws.parent["ValidationData"]
            
            # Find next available column
            col = 1
            while validation_sheet.cell(row=1, column=col).value:
                col += 1
            
            # Add options to validation sheet
            for i, option in enumerate(options, 1):
                validation_sheet.cell(row=i, column=col, value=option)
            
            # Create named range
            range_name = f"Options_{field_def['field_id']}"
            validation_sheet.parent.define_name(
                range_name, 
                f"ValidationData.{validation_sheet.cell(row=1, column=col).coordinate}:{validation_sheet.cell(row=len(options), column=col).coordinate}"
            )
            
            dv = DataValidation(type="list", formula1=range_name)
        
        dv.error = "Please select from the dropdown list"
        dv.errorTitle = "Invalid Selection"
        dv.prompt = "Click the dropdown arrow to select an option"
        dv.promptTitle = "Select Option"
        
        if field_def['field_type'] == 'multi-select':
            dv.error = "For multiple selections, separate values with commas"
            dv.prompt = "You can enter multiple values separated by commas, or select from dropdown"
        
        ws.add_data_validation(dv)
        dv.add(cell)

def add_text_validation(ws, cell, field_def):
    """Add text input validation."""
    field_type = field_def['field_type']
    required = field_def['required'].lower() == 'true'
    
    if required:
        # Add formula to check if cell is not empty
        dv = DataValidation(type="custom", formula1=f'LEN({cell.coordinate})>0')
        dv.error = "This field is required and cannot be empty"
        dv.errorTitle = "Required Field"
        dv.prompt = "Please enter the required information"
        dv.promptTitle = "Required Field"
        ws.add_data_validation(dv)
        dv.add(cell)
    
    # Set text wrapping for textarea
    if field_type == 'textarea':
        cell.alignment = Alignment(wrap_text=True, vertical='top')

def add_checkbox_options(ws, cell, field_def, data, start_row):
    """Add checkbox-style options using multiple cells."""
    options = get_field_options(field_def, data)
    
    if not options:
        return
    
    # For checkboxes, we'll use a different approach
    # Create a merged cell with instructions
    cell.value = "Check applicable options below:"
    
    # Add checkbox options in subsequent rows
    for i, option in enumerate(options[:8]):  # Limit to 8 options
        option_row = start_row + i + 1
        
        # Checkbox column (using a symbol)
        checkbox_cell = ws[f'B{option_row}']
        checkbox_cell.value = f"☐ {option}"
        checkbox_cell.style = "input_style"
        
        # Add data validation for checkbox
        dv = DataValidation(type="list", formula1='"☐,☑"')
        dv.prompt = "Click to check/uncheck this option"
        ws.add_data_validation(dv)
        dv.add(checkbox_cell)

def get_field_options(field_def, data):
    """Extract options for dropdown and multi-select fields."""
    data_source = field_def['data_source']
    options_type = field_def['options_type']
    
    options = []
    
    if data_source != 'none' and data_source in data:
        source_data = data[data_source]
        
        if options_type == 'dict_keys':
            options = list(source_data.keys())
        elif options_type == 'list':
            options = list(source_data)
        elif options_type in ['dict_multiple']:
            options = list(source_data.keys())
        elif options_type == 'dict_with_extra':
            options = list(source_data.keys())
            options.extend(["Open Source", "Registration Required", "Proprietary"])
        elif options_type == 'list_with_na':
            options = ["Not applicable"] + list(source_data)
    
    # Clean options for Excel (remove problematic characters)
    cleaned_options = []
    for option in options:
        # Remove quotes and limit length
        clean_option = str(option).replace('"', '').replace("'", "")[:100]
        if clean_option:
            cleaned_options.append(clean_option)
    
    return cleaned_options

def get_filling_instructions(field_type, required, field_def, data):
    """Generate filling instructions based on field type."""
    instructions = []
    
    if field_type == 'input':
        instructions.append("Enter text in the cell")
    elif field_type == 'textarea':
        instructions.append("Enter multiple lines of text (press Alt+Enter for new lines)")
    elif field_type == 'dropdown':
        instructions.append("Click the dropdown arrow to select an option")
    elif field_type == 'multi-select':
        instructions.append("Select from dropdown or enter multiple values separated by commas")
    elif field_type == 'checkboxes':
        instructions.append("Select checkboxes for applicable options")
    
    if required:
        instructions.append("This field is required")
    else:
        instructions.append("This field is optional")
    
    return ". ".join(instructions) + "."

def create_summary_sheet(wb, config, total_fields):
    """Create a summary sheet with completion status."""
    ws = wb.create_sheet("Summary & Validation")
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    
    # Title
    ws.merge_cells('A1:B2')
    ws['A1'] = "Form Completion Summary"
    ws['A1'].style = "title_style"
    ws.row_dimensions[1].height = 30
    
    # Summary information
    row = 4
    summary_data = [
        ("Template Name:", config['name']),
        ("Total Fields:", total_fields),
        ("Required Fields:", "Check validation messages"),
        ("Status:", "=IF(COUNTA('Form Fields'!C:C)>1,\"In Progress\",\"Not Started\")")
    ]
    
    for label, value in summary_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].style = "field_label_style"
        if isinstance(value, str) and value.startswith('='):
            ws[f'B{row}'].value = value  # Excel formula
        else:
            ws[f'B{row}'] = str(value)
        ws[f'B{row}'].style = "input_style"
        row += 1
    
    # Instructions
    row += 2
    instructions = [
        "Next Steps:",
        "• Complete all required fields in the 'Form Fields' tab",
        "• Review your responses for accuracy",
        "• Save the file when complete",
        "• Submit through the appropriate channel",
        "",
        "Validation:",
        "• Required fields will show error messages if empty",
        "• Dropdown fields must use valid options",
        "• Hover over cells for help text"
    ]
    
    for instruction in instructions:
        ws[f'A{row}'] = instruction
        if instruction.endswith(':'):
            ws[f'A{row}'].style = "header_style"
        else:
            ws[f'A{row}'].style = "desc_style"
        row += 1
    
    return ws

def protect_workbook(wb):
    """Protect the workbook while keeping input cells editable."""
    for ws in wb.worksheets:
        if ws.title == "Form Fields":
            # Unlock input cells (column C)
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    if cell.value is not None or cell.style == "input_style":
                        cell.protection = cell.protection.copy(locked=False)
            
            # Protect the sheet
            ws.protection.sheet = True
            ws.protection.password = None  # No password for easier use
            ws.protection.formatCells = False
            ws.protection.formatColumns = False
            ws.protection.formatRows = False
        elif ws.title != "ValidationData":
            # Protect other sheets completely
            ws.protection.sheet = True

def generate_excel_workbook(config, fields, data):
    """Generate complete Excel workbook with interactive forms."""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Set up styles
    setup_workbook_styles(wb)
    
    # Create sheets
    create_overview_sheet(wb, config, fields)
    create_form_sheet(wb, config, fields, data)
    
    # Count actual form fields
    field_count = len([f for f in fields if f['field_type'] != 'markdown'])
    create_summary_sheet(wb, config, field_count)
    
    # Protect workbook
    protect_workbook(wb)
    
    return wb

def process_template_pair(template_name, csv_file, py_file, output_dir):
    """Process a single template pair and generate Excel workbook."""
    
    print(f"Processing {template_name}...")
    
    try:
        config, data = load_template_data(py_file)
        if not config:
            print(f"    No TEMPLATE_CONFIG found")
            return False
        
        print(f"    Loaded: {config['name']}")
        
        fields = load_csv_fields(csv_file)
        print(f"    Fields: {len(fields)}")
        
        wb = generate_excel_workbook(config, fields, data)
        
        # Save the workbook
        output_file = output_dir / f"{template_name}_form.xlsx"
        wb.save(str(output_file))
        
        print(f"    Generated {template_name}_form.xlsx")
        return True
        
    except Exception as e:
        print(f"    Error: {e}")
        failed_templates.append(f"{template_name}: {str(e)}")
        return False

def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description='Generate Excel interactive forms for issue templates')
    parser.add_argument('-t', '--template-dir', type=Path, help='Template directory')
    parser.add_argument('-o', '--output-dir', type=Path, help='Output directory')
    parser.add_argument('--template', type=str, help='Generate form for specific template')
    return parser.parse_args()

def main():
    """Main function."""
    args = parse_arguments()
    
    print("Excel Interactive Template Generator")
    print("Generating interactive forms for issue templates")
    
    script_dir = Path.cwd()
    template_dir = args.template_dir or script_dir / ".github" / "GEN_ISSUE_TEMPLATE"
    output_dir = args.output_dir or script_dir / ".github" / "template_forms_xlsx"
    
    print(f"Template dir: {template_dir}")
    print(f"Output dir: {output_dir}")
    
    if not template_dir.exists():
        print(f"Template directory not found: {template_dir}")
        return False
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    csv_files = list(template_dir.glob('*.csv'))
    
    if args.template:
        csv_files = [f for f in csv_files if f.stem == args.template]
    
    if not csv_files:
        print("No CSV files found")
        return False
    
    print(f"Processing {len(csv_files)} templates")
    
    success_count = 0
    for csv_file in csv_files:
        template_name = csv_file.stem
        py_file = template_dir / f"{template_name}.py"
        
        if py_file.exists():
            if process_template_pair(template_name, csv_file, py_file, output_dir):
                success_count += 1
        else:
            print(f"    Warning: No Python config file found for {template_name}")
    
    print(f"Results: {success_count}/{len(csv_files)} successful")
    
    if failed_templates:
        print("\\nFailed templates:")
        for failure in failed_templates:
            print(f"  - {failure}")
    
    print("\\nNote: Generated Excel files include:")
    print("- Interactive dropdown menus with data validation")
    print("- Protected sheets with unlocked input areas")
    print("- Help comments on hover")
    print("- Form completion summary")
    
    return success_count > 0

if __name__ == '__main__':
    try:
        success = main()
        sys.exit(0 if success else 1)
    except ImportError as e:
        if 'openpyxl' in str(e):
            print("Error: openpyxl package not found. Please install it:")
            print("pip install openpyxl")
            sys.exit(1)
        else:
            raise
