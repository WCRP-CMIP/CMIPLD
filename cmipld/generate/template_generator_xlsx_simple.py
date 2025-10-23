#!/usr/bin/env python3
"""
Excel Template Generator - Simplified and Fixed

pip install openpyxl
"""

import csv
import sys
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

failed_templates = []

def load_template_data(py_file):
    """Load configuration and data from Python file."""
    namespace = {}
    try:
        with open(py_file, 'r', encoding='utf-8') as f:
            exec(f.read(), namespace)
        config = namespace.get('TEMPLATE_CONFIG', {})
        data = namespace.get('DATA', {})
        return config, data
    except Exception as e:
        print(f"    Error loading template data: {e}")
        return {}, {}

def load_csv_fields(csv_file):
    """Load field definitions from CSV."""
    fields = []
    try:
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                fields.append(row)
        fields.sort(key=lambda x: int(x.get('field_order', 0)))
        return fields
    except Exception as e:
        print(f"    Error loading CSV fields: {e}")
        return []

def get_field_options(field_def, data):
    """Extract options for dropdown and multi-select fields."""
    try:
        data_source = field_def.get('data_source', 'none')
        options_type = field_def.get('options_type', '')
        
        options = []
        
        if data_source != 'none' and data_source in data and data[data_source] is not None:
            source_data = data[data_source]
            
            if options_type == 'dict_keys' and hasattr(source_data, 'keys'):
                options = list(source_data.keys())
            elif options_type == 'list' and hasattr(source_data, '__iter__'):
                options = list(source_data)
            elif options_type == 'dict_multiple' and hasattr(source_data, 'keys'):
                options = list(source_data.keys())
            elif options_type == 'dict_with_extra' and hasattr(source_data, 'keys'):
                options = list(source_data.keys())
                options.extend(["Open Source", "Registration Required", "Proprietary"])
            elif options_type == 'list_with_na':
                options = ["Not applicable"]
                if hasattr(source_data, '__iter__'):
                    options.extend(list(source_data))
        
        # Clean options for Excel
        cleaned_options = []
        for option in options[:30]:  # Limit to 30 options to avoid Excel issues
            if option is not None:
                clean_option = str(option).replace('"', '').replace("'", "")[:100]
                if clean_option:
                    cleaned_options.append(clean_option)
        
        return cleaned_options
    except Exception as e:
        print(f"    Warning: Could not get options for field {field_def.get('field_id', 'unknown')}: {e}")
        return []

def add_dropdown_validation(ws, cell, options):
    """Add simple dropdown validation."""
    try:
        if options and len(options) <= 10:  # Only for small option lists
            options_str = ",".join([f'"{opt}"' for opt in options])
            dv = DataValidation(type="list", formula1=options_str)
            dv.prompt = "Click the dropdown arrow to select an option"
            dv.promptTitle = "Select Option"
            ws.add_data_validation(dv)
            dv.add(cell)
        elif options:
            # Too many options, just add a comment
            comment_text = "Options: " + ", ".join(options[:5])
            if len(options) > 5:
                comment_text += f" ... ({len(options)} total options)"
            cell.comment = Comment(comment_text, "System")
    except Exception as e:
        print(f"    Warning: Could not add dropdown: {e}")

def create_form_sheet(wb, config, fields, data):
    """Create the main form sheet."""
    try:
        ws = wb.create_sheet("Form Fields")
        
        # Set column widths
        ws.column_dimensions['A'].width = 3   
        ws.column_dimensions['B'].width = 35  
        ws.column_dimensions['C'].width = 50  
        ws.column_dimensions['D'].width = 30  
        
        # Header row
        headers = ["#", "Field", "Your Response", "Information"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
        
        current_row = 2
        field_number = 1
        
        for field_def in fields:
            try:
                field_type = field_def.get('field_type', '')
                field_id = field_def.get('field_id', '')
                label = field_def.get('label', '')
                description = field_def.get('description', '')
                required = field_def.get('required', '').lower() == 'true'
                
                # Skip markdown fields
                if field_type == 'markdown':
                    if description:
                        ws.merge_cells(f'A{current_row}:D{current_row}')
                        cell = ws[f'A{current_row}']
                        formatted_desc = description.replace('\\n', ' ')
                        cell.value = f"Information: {formatted_desc}"
                        current_row += 2
                    continue
                
                # Field number
                ws[f'A{current_row}'] = field_number
                
                # Field label
                label_text = label
                if required:
                    label_text += " *"
                ws[f'B{current_row}'] = label_text
                
                # Input cell
                input_cell = ws[f'C{current_row}']
                input_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                
                # Field information
                info_text = f"Type: {field_type.title()}"
                if description:
                    formatted_desc = description.replace('\\n', ' ')
                    info_text += f"\\nDescription: {formatted_desc}"
                ws[f'D{current_row}'] = info_text
                
                # Add dropdown for dropdown/multi-select fields
                if field_type in ['dropdown', 'multi-select']:
                    options = get_field_options(field_def, data)
                    if options:
                        add_dropdown_validation(ws, input_cell, options)
                        # Add options to info
                        current_info = ws[f'D{current_row}'].value or ""
                        options_text = "\\nOptions: " + ", ".join(options[:5])
                        if len(options) > 5:
                            options_text += f" ... ({len(options)} total)"
                        ws[f'D{current_row}'].value = current_info + options_text
                
                # Add help comment
                if description:
                    help_text = description.replace('\\n', '\\n')
                    if field_type in ['input', 'textarea']:
                        help_text += f"\\n\\nEnter text in the cell."
                    elif field_type == 'dropdown':
                        help_text += f"\\n\\nSelect from dropdown options."
                    if required:
                        help_text += " This field is required."
                    input_cell.comment = Comment(help_text[:1000], "System")  # Limit comment length
                
                current_row += 2
                field_number += 1
                
            except Exception as e:
                print(f"    Warning: Error processing field {field_def.get('field_id', 'unknown')}: {e}")
                current_row += 1
        
        return ws
    except Exception as e:
        print(f"    Error creating form sheet: {e}")
        return None

def generate_excel_workbook(config, fields, data):
    """Generate Excel workbook."""
    try:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create title sheet
        try:
            title_ws = wb.create_sheet("Overview", 0)
            title_ws['A1'] = config.get('name', 'Template Form')
            title_ws['A1'].font = Font(size=18, bold=True)
            
            title_ws['A3'] = f"Description: {config.get('description', 'Form generated from template')}"
            title_ws['A4'] = f"Total Fields: {len([f for f in fields if f.get('field_type') != 'markdown'])}"
            title_ws['A5'] = "Navigate to the 'Form Fields' tab to complete the form."
        except Exception as e:
            print(f"    Warning: Could not create overview sheet: {e}")
        
        # Create form sheet
        form_ws = create_form_sheet(wb, config, fields, data)
        if not form_ws:
            raise Exception("Could not create form sheet")
        
        return wb
    except Exception as e:
        print(f"    Error generating workbook: {e}")
        raise

def process_template_pair(template_name, csv_file, py_file, output_dir):
    """Process a single template pair."""
    print(f"Processing {template_name}...")
    
    try:
        config, data = load_template_data(py_file)
        if not config:
            config = {'name': f'{template_name} Template', 'description': 'Generated template form'}
        
        print(f"    Loaded: {config.get('name', 'Unknown')}")
        
        fields = load_csv_fields(csv_file)
        if not fields:
            print(f"    No fields found")
            return False
        
        print(f"    Fields: {len(fields)}")
        
        wb = generate_excel_workbook(config, fields, data)
        
        output_file = output_dir / f"{template_name}_form.xlsx"
        wb.save(str(output_file))
        
        print(f"    Generated {template_name}_form.xlsx")
        return True
        
    except Exception as e:
        print(f"    Error: {e}")
        failed_templates.append(f"{template_name}: {str(e)}")
        return False

def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description='Generate Excel forms for issue templates')
    parser.add_argument('-t', '--template-dir', type=Path, help='Template directory')
    parser.add_argument('-o', '--output-dir', type=Path, help='Output directory')
    parser.add_argument('--template', type=str, help='Generate specific template')
    return parser.parse_args()

def main():
    """Main function."""
    args = parse_arguments()
    
    print("Excel Interactive Template Generator - Simplified")
    
    script_dir = Path.cwd()
    template_dir = args.template_dir or script_dir / ".github" / "GEN_ISSUE_TEMPLATE"
    output_dir = args.output_dir or script_dir / ".github" / "template_forms_xlsx"
    
    print(f"Template dir: {template_dir}")
    print(f"Output dir: {output_dir}")
    
    if not template_dir.exists():
        print(f"Template directory not found: {template_dir}")
        return False
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    csv_files = list(template_dir.glob('*.csv'))
    
    if args.template:
        csv_files = [f for f in csv_files if f.stem == args.template]
    
    if not csv_files:
        print("No CSV files found")
        return False
    
    print(f"Processing {len(csv_files)} templates")
    
    success_count = 0
    for csv_file in csv_files:
        template_name = csv_file.stem
        py_file = template_dir / f"{template_name}.py"
        
        if py_file.exists():
            if process_template_pair(template_name, csv_file, py_file, output_dir):
                success_count += 1
        else:
            print(f"    Warning: No Python config file found for {template_name}")
    
    print(f"Results: {success_count}/{len(csv_files)} successful")
    
    if failed_templates:
        print("\\nFailed templates:")
        for failure in failed_templates:
            print(f"  - {failure}")
    
    return success_count > 0

if __name__ == '__main__':
    try:
        success = main()
        sys.exit(0 if success else 1)
    except ImportError as e:
        if 'openpyxl' in str(e):
            print("Error: openpyxl package not found. Please install it:")
            print("pip install openpyxl")
            sys.exit(1)
        else:
            raise
